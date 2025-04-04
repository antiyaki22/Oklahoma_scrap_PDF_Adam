import asyncio
import os
from datetime import datetime
import re
import subprocess
from sdk.extract_text_info_from_pdf import ExtractTextInfoFromPDF
from sdk.extract_text_info_with_char_bounds_from_pdf import ExtractTextInfoWithCharBoundsFromPDF
import json
import spacy
from spacy.matcher import Matcher
import zipfile
import usaddress
import phonenumbers
import pandas as pd
from openpyxl import load_workbook, Workbook
from playwright.async_api import async_playwright

TARGET_URL = "https://www.okcc.online/index.php"
CSV_FILE = "result.csv"
XLSX_FILE = "result.xlsx"

TABLE_HEADER_SELECTOR = "#rod-table thead tr th"
TABLE_ROW_SELECTOR = "#rodinitialbody tr"
TABLE_CELL_SELECTOR = "td"

nlp = spacy.load("en_core_web_sm")
months = 3

def extract_company_name(text):
    doc = nlp(text)
    matcher = Matcher(nlp.vocab)

    company_suffixes = ["INC", "LLC", "CORP", "CORPORATION", "GROUP", "ENTERPRISES", "HOLDINGS", "DBA", "CO",
                        "LIMITED", "PARTNERSHIP", "ASSOCIATION", "COMPANY"]

    company_patterns = [
        [{"IS_ALPHA": True, "OP": "+"}, {"TEXT": {"in": company_suffixes}}],  
        [{"IS_ALPHA": True, "OP": "+"}, {"IS_PUNCT": True}, {"IS_ALPHA": True, "OP": "+"}, {"TEXT": {"in": company_suffixes}}],
        [{"TEXT": {"in": ["DBA"]}}, {"IS_ALPHA": True, "OP": "+"}, {"IS_ALPHA": True, "OP": "+"}],
    ]

    for pattern in company_patterns:
        matcher.add("COMPANY_NAME_PATTERN", [pattern])

    matches = matcher(doc)
    company_names = []

    for match_id, start, end in matches:
        span = doc[start:end]
        
        if span.text.strip() in company_suffixes and start > 0:
            span = doc[start - 1:end] 

        elif span.text.split()[-1] in company_suffixes and start > 0:
            prev_token = doc[start - 1]
            if prev_token.is_alpha:
                span = doc[start - 1:end]

        company_names.append(span.text.strip())

    company_names = [name for name in company_names if not re.search(r'\d{1,5}\s\w+(\s\w+)*', name)]

    if company_names:
        company_names.sort(key=len, reverse=True)
        return company_names[0]

    match = re.search(r"([A-Za-z\s]+(?:,\s[A-Za-z\s]+)*\s*,?\s*(?:LLC|INC|CORP|CORPORATION|GROUP|ENTERPRISES|HOLDINGS|DBA|CO|LIMITED|PARTNERSHIP|ASSOCIATION)(?:\s*\([^)]+\))?)", text)
    
    if match:
        return match.group(0).strip()

    for ent in doc.ents:
        if ent.label_ == "PERSON" and ent.text not in company_names:
            company_names.append(ent.text.strip())

    if company_names:
        company_names.sort(key=len, reverse=True)
        return company_names[0]

    return None

def extract_phone_number(text):
    numbers = [match.number for match in phonenumbers.PhoneNumberMatcher(text, "US")]
    if numbers:
        phone_number = phonenumbers.format_number(numbers[0], phonenumbers.PhoneNumberFormat.INTERNATIONAL)
        return phone_number.replace(" ", "-")
    return None

def ensure_playwright_browsers():
    try:
        subprocess.run(["playwright", "install", "--with-deps"], check=True, shell=True)
    except Exception as e:
        print(f"Error installing Playwright: {e}")

def clear_downloads_output_folder(download_path, output_path):
    if os.path.exists(download_path):  
        for filename in os.listdir(download_path):
            file_path = os.path.join(download_path, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path) 
                    print(f"Removed: {file_path}")
                elif os.path.isdir(file_path):
                    os.rmdir(file_path) 
                    print(f"Removed directory: {file_path}")
            except Exception as e:
                print(f"Error removing {file_path}: {e}")
    else:
        print(f"Directory {download_path} does not exist.")
    if os.path.exists(output_path):  
        for filename in os.listdir(output_path):
            file_path = os.path.join(output_path, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path) 
                    print(f"Removed: {file_path}")
                elif os.path.isdir(file_path):
                    os.rmdir(file_path) 
                    print(f"Removed directory: {file_path}")
            except Exception as e:
                print(f"Error removing {file_path}: {e}")
    else:
        print(f"Directory {output_path} does not exist.")

def clear_xlsx_file():
    if os.path.isfile(XLSX_FILE):
        wb = Workbook()  
        wb.save(XLSX_FILE)

def fix_misplaced_decimal(amount):
    """
    Fix misplaced decimal formatting like "22.692.92" -> "22692.92"
    """
    amount = amount.replace(" ", "").replace(",", "")  
    parts = amount.split(".")
    
    if len(parts) > 2:  
        amount = parts[0] + parts[1] + "." + parts[-1]  
    
    return amount

def extract_dollar_amount(json_file_path):
    with open(json_file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    patterns = [
        r"of \$\s?([\d,]+\.\d{1,2})",
        r"\(\$\s?([\d,]+\.\d{1,2})\)",
        r"\$\s?([\d,]+\.\d{1,2}) due",
        r"is \$\s?([\d,]+\.\d{1,2})",
        r"total \$\s?([\d,]+\.\d{1,2})",
        r"is\$\s?([\d,]+\.\d{1,2})",
        r"of\$\s?([\d,]+\.\d{1,2})",
        r"j\$([\d,]+\.\d{1,2})",
        r"j \$([\d,]+\.\d{1,2})"
    ]

    all_amounts = []
    elements = data.get("elements", [])
    
    for i, element in enumerate(elements):
        text = element.get("Text", "")

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return fix_misplaced_decimal(match.group(1))

        dollar_matches = re.findall(r"\$\s?([\d,]+\.\d{1,2})", text)
        all_amounts.extend(fix_misplaced_decimal(m) for m in dollar_matches)

        if "Principal amount of claim:" in text:
            for j in range(1, 3):
                if i + j < len(elements):
                    next_text = elements[i + j].get("Text", "")
                    next_dollar_matches = re.findall(r"\$\s?([\d,]+\.\d{1,2})", next_text)
                    if next_dollar_matches:
                        return fix_misplaced_decimal(next_dollar_matches[0]) 

    if all_amounts:
        return max(all_amounts, key=lambda x: float(x.replace(",", "")))

    return "0"

def extract_full_name(json_file_path):
    with open(json_file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    full_names = []  
    priority_name = None  

    for element in data.get("elements", []):
        text = element.get("Text", "")
        doc = nlp(text)

        for ent in doc.ents:
            if ent.label_ == "PERSON":
                name_parts = ent.text.split()

                if len(name_parts) > 1:  
                    full_names.append(ent.text)
                
                match = re.search(r"against\s+" + re.escape(ent.text), text, re.IGNORECASE)
                if match:
                    priority_name = ent.text  

    return priority_name if priority_name else (full_names[0] if full_names else None)

def extract_address(text):
    try:
        if not text:
            return None, None, None, None

        def clean_text(text):
            return re.sub(r'[^\x00-\x7F]+', ' ', text).strip()

        text = clean_text(text)
        text = re.sub(r'\s+', ' ', text)

        print("Cleaned Text:", text)  

        try:
            parsed_address = usaddress.parse(text)
            print("usaddress.parse output:", parsed_address) 

            address_number = None
            street_name = []
            street_name_post_type = None
            place_name = None
            state_name = None
            zip_code = None

            for component, label in parsed_address:
                if label == "AddressNumber" and not address_number:
                    address_number = component
                elif label == "StreetName" and not street_name:
                    street_name.append(component)
                elif label == "StreetNamePostType" and not street_name_post_type:
                    street_name_post_type = component
                elif label == "PlaceName" and not place_name:
                    place_name = component.strip(",")
                elif label == "StateName" and not state_name:
                    state_name = component
                elif label == "ZipCode" and not zip_code:
                    zip_code = component

            if address_number and street_name and street_name_post_type:
                best_address = " ".join([address_number] + street_name + [street_name_post_type])
            elif street_name:
                best_address = " ".join([address_number] + street_name) if address_number else " ".join(street_name)

            best_city = place_name
            best_state = state_name
            best_zipcode = zip_code

            print("Final Address:", best_address, best_city, best_state, best_zipcode)

        except usaddress.RepeatedLabelError:
            print("usaddress.parse failed")

        return best_address, best_city, best_state, best_zipcode

    except Exception as e:
        print(f"Error in extract_address: {e}")
        return None, None, None, None
        
def get_merged_text(file_path: str) -> str:
    with open(file_path, 'r') as file:
        json_data = json.load(file)

    merged_text = ""
    for element in json_data.get("elements", []):
        if "Text" in element:
            merged_text += element["Text"] + " "
    
    return merged_text.strip()

def get_claimant(text):
    claimant_match = re.search(r'claimant:\s*(\S+(?:\s+\S+){0,19})', text, re.IGNORECASE | re.DOTALL)
    if claimant_match:
        claimant_text = claimant_match.group(1).strip()
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print(f"claimant: {claimant_text}")
        claimant_name = extract_company_name(claimant_text)
        if claimant_name:
            return claimant_name

    claims_match = re.search(r'(\S+(?:\s+\S+){0,19})\s+\b(?:claims|against|upon)\b', text, re.IGNORECASE | re.DOTALL)
    if claims_match:
        claimant_text = claims_match.group(1).strip()
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print(f"claimant: {claimant_text}")
        claimant_name = extract_company_name(claimant_text)
        if claimant_name:
            return claimant_name

    return None

def get_contractor(text):
    contractor_match = re.search(r'\b(?:Contractor|Customer|claims|against|upon):?\s*(\S+(?:\s+\S+){0,29})', text, re.IGNORECASE | re.DOTALL)
    if contractor_match:
        contractor_text = contractor_match.group(1).strip()
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print(f"contractor: {contractor_text}")
        contractor_name = extract_company_name(contractor_text)
        if contractor_name:
            return contractor_name

    return None

def get_owner(text):
    owner_match = re.search(r'\b(?:Owner|Owners|owned by|owned)\b:?\s*(\S+(?:\s+\S+){0,29})', text, re.IGNORECASE | re.DOTALL)
    if owner_match:
        owner_text = owner_match.group(1).strip()
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print(f"owner: {owner_text}")
        owner_name = extract_company_name(owner_text)
        if owner_name:
            return owner_name

    return None

def get_property_address(text):
    property_match = re.search(r'\b(?:property:|contract:|notice to:|prepared by:|following:)\b:?\s*(\S+(?:\s+\S+){0,29})', text, re.IGNORECASE | re.DOTALL)
    if property_match:
        property_text = property_match.group(1).strip()
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print(f"property: {property_text}")
        address, city, state, zip = extract_address(property_text)
        if address or city or state or zip:
            return address, city, state, zip

    property_match = re.search(r'\b(?:against|upon)\b:?\s*(\S+(?:\s+\S+){0,49})', text, re.IGNORECASE | re.DOTALL)
    if property_match:
        property_text = property_match.group(1).strip()
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print(f"property: {property_text}")
        address, city, state, zip = extract_address(property_text)
        if address or city or state or zip:
            return address, city, state, zip

    return None, None, None, None

def get_claimant_phone(text):
    claimant_match = re.search(r'claimant:\s*(\S+(?:\s+\S+){0,29})', text, re.IGNORECASE | re.DOTALL)
    if claimant_match:
        claimant_text = claimant_match.group(1)
        phone = extract_phone_number(claimant_text)
        if phone:
            return phone

    claims_match = re.search(r'(\S+(?:\s+\S+){0,29})\s+\b(?:claims|against|upon)\b', text, re.IGNORECASE | re.DOTALL)
    if claims_match:
        claimant_text = claims_match.group(1)
        phone = extract_phone_number(claimant_text)
        if phone:
            return phone

    return None

def unzip_file(zip_file_path, output_folder):
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(output_folder)

def remove_zip_file(zip_file_path):
    try:
        os.remove(zip_file_path)
    except Exception as e:
        print(f"Error removing zip file: {e}")

async def set_table_headers(page) -> list:
    header_titles = []
    header_titles.append("File")
    header_titles.append("Instrument Number")
    header_titles.append("Type")
    header_titles.append("Date Recorded")
    header_titles.append("Book")
    header_titles.append("Page")
    header_titles.append("Claimant")
    header_titles.append("Contractor")
    header_titles.append("Owner")
    header_titles.append("Property Address")
    header_titles.append("Property City")
    header_titles.append("Property State")
    header_titles.append("Property Zip")
    header_titles.append("Dollar Amount")
    header_titles.append("Phone Number")
    return header_titles

async def download_pdf(page, key: str, docid: str) -> bool:
    pdf_url = None
    def response_handler(response):
        nonlocal pdf_url
        if response.url.startswith("https://www.okcc.online/document.php") and response.headers.get('content-type', '').startswith('application/pdf'):
            pdf_url = response.url

    page.on('response', response_handler)

    await page.evaluate(f'OpenP("{key}", document.body, "{docid}");')
    await asyncio.sleep(5)

    download_path = os.path.join(os.getcwd(), 'downloads')
    os.makedirs(download_path, exist_ok=True)

    if pdf_url:
        response = await page.request.get(pdf_url)

        if response.ok:
            pdf_content = await response.body()
            pdf_path = os.path.join(download_path, f"{docid}.pdf")
            
            with open(pdf_path, 'wb') as pdf_file:
                pdf_file.write(pdf_content)     
        else:
            print("❌ Failed to fetch PDF.")
        
        await page.click(".pdf-close")
        await asyncio.sleep(2)
        return True
    else:
        return False

def remove_watermark(wm_text, inputFile, outputFile):
    from PyPDF4 import PdfFileReader, PdfFileWriter
    from PyPDF4.pdf import ContentStream
    from PyPDF4.generic import TextStringObject, NameObject
    from PyPDF4.utils import b_
    
    with open(inputFile, "rb") as f:
        source = PdfFileReader(f, "rb")
        output = PdfFileWriter()

        for page in range(source.getNumPages()):
            page = source.getPage(page)
            content_object = page["/Contents"].getObject()
            content = ContentStream(content_object, source)

            for operands, operator in content.operations:
                if operator == b_("Tj"):
                    text = operands[0]

                    if isinstance(text, str) and text.startswith(wm_text):
                        operands[0] = TextStringObject('')

            page.__setitem__(NameObject('/Contents'), content)
            output.addPage(page)

        with open(outputFile, "wb") as outputStream:
            output.write(outputStream)

async def process_pdf(docid: str) -> tuple:
    input_pdf_path = f"downloads/{docid}.pdf"
    pdf_filename = os.path.splitext(os.path.basename(input_pdf_path))[0]
    ExtractTextInfoWithCharBoundsFromPDF(input_pdf_path)
    
    output_folder = "output/ExtractTextInfoWithCharBoundsFromPDF"
    time_stamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    zip_file_path = f"{output_folder}/extract{time_stamp}.zip"
        
    unzip_file(zip_file_path, output_folder)
    remove_zip_file(zip_file_path)

    json_file_path = f"{output_folder}/structuredData.json"
    renamed_json_path = f"{output_folder}/{pdf_filename}.json"

    os.rename(json_file_path, renamed_json_path)    

    full_text = get_merged_text(renamed_json_path)

    claimant = get_claimant(full_text)
    contractor = get_contractor(full_text)
    owner = get_owner(full_text)
    address, city, state, zipcode = get_property_address(full_text)
    dollar_amount = f"${extract_dollar_amount(renamed_json_path)}"
    phone_number = get_claimant_phone(full_text)

    info: dict[str, any] = {
        "claimant": claimant,
        "contractor": contractor,
        "owner": owner,
        "address": address,
        "city": city,
        "state": state,
        "zipcode": zipcode,
        "dollar": dollar_amount,
        "phone": phone_number,
    }

    print (f"info: {info}")
    return info

async def scrape_table(page, headers):
    rows = await page.query_selector_all(TABLE_ROW_SELECTOR)

    for row in rows:
        cells = await row.query_selector_all(TABLE_CELL_SELECTOR)
        cell_values = [((await cell.text_content()) or "").strip() or "N/A" for cell in cells]

        pdf_html_element = await cells[0].query_selector("div > button:first-of-type")
        pdf_html = await pdf_html_element.evaluate("element => element.outerHTML")

        match = re.search(r"OpenP\('([^']+)',this,'([^']+)'\)", str(pdf_html))

        if match:
            instrument_number = match.group(1) 
            doc_id = match.group(2) 
        else:
            print("Document not found!")

        downloaded = await download_pdf(page, key=instrument_number, docid=doc_id)
        if downloaded:
            input_path = f"downloads/{doc_id}.pdf"
            temp_output_path = f"downloads/{doc_id}_no_watermark.pdf"

            remove_watermark("UNOFFICIAL", input_path, temp_output_path)

            if os.path.exists(temp_output_path):
                os.remove(input_path) 
                os.rename(temp_output_path, input_path) 
                print(f"Successfully replaced {input_path} with watermark-free version.")
            else:
                print("Error: Watermark removal failed, new file not created.")

            cell_values[0] = f"{doc_id}.pdf"
            print (f"cell values 0: ", cell_values)

            info = await process_pdf(docid=doc_id)
            if cell_values[6] == "N/A":
                cell_values[6] = info["claimant"]
            if cell_values[7] == "N/A":
                cell_values[7] = info["contractor"]
            # if cell_values[8] == "N/A":
            cell_values[8] = info["owner"]
            cell_values[9] = info["address"]
            cell_values.append(info["city"])
            cell_values.append(info["state"])
            cell_values.append(info["zipcode"])
            cell_values.append(info["dollar"])
            cell_values.append(info["phone"])

        save_to_xlsx([cell_values], headers=None, append=True)
        await asyncio.sleep(2)

def save_to_xlsx(data, headers, append=True):
    if append and os.path.isfile(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
        start_row = ws.max_row 
    else:
        wb = Workbook()  
        ws = wb.active
        start_row = 1  

    if headers and start_row == 1:
        ws.append(headers) 

    if data:
        for row in data:
            ws.append(row)

    wb.save(XLSX_FILE)

    print(f"Updated {XLSX_FILE} with new data: {data if data else 'No data'} and headers: {headers if headers else 'No headers'}")

async def main():    
    clear_xlsx_file()

    download_path = os.path.join(os.getcwd(), 'downloads')
    output_path = os.path.join(os.getcwd(), 'output/ExtractTextInfoWithCharBoundsFromPDF')
    os.makedirs(download_path, exist_ok=True)
    clear_downloads_output_folder(download_path, output_path)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()

        await page.goto(TARGET_URL, timeout=60000)

        await page.click("div#areastyle > div.col-md-4:first-of-type ul.text-start i.fa-file-magnifying-glass")
        await page.wait_for_selector("input#rodDocTypeTxt")
        await page.fill("input#rodDocTypeTxt", "ml")
        await page.click("text='ML - MECHANIC LIEN'")
        await page.click("#date_range_rod_type")

        today = str(datetime.today().day)

        await page.click('#drwrapper-rod-type #rodDateFromTxt')
        await asyncio.sleep(1)
        
        for i in range(months):
            await page.click('div.flatpickr-calendar.open .flatpickr-months .flatpickr-prev-month svg')
        
        ### From Date Click ###
        dayContainer_from = page.locator('div.flatpickr-calendar.open .flatpickr-innerContainer .dayContainer')
        all_spans = dayContainer_from.locator('span')
        from_date = None

        for index in range(await all_spans.count()):  
            span_element = all_spans.nth(index) 
            text_content = await span_element.inner_text() 
            class_attribute = await span_element.get_attribute("class") 

            if text_content == today and ("prevMonthDay" not in (class_attribute or "")) and ("nextMonthDay" not in (class_attribute or "")):
                from_date = span_element
                break  

        if from_date:
            await from_date.click()
        else:
            print("No valid date found!")
        ###################

        await page.click('#drwrapper-rod-type #rodToDateTxt')
        await asyncio.sleep(1)

        ### To Date Click ###
        dayContainer_to = page.locator('div.flatpickr-calendar.open .flatpickr-innerContainer .dayContainer')
        all_spans = dayContainer_to.locator('span')
        to_date = None

        for index in range(await all_spans.count()):  
            span_element = all_spans.nth(index) 
            text_content = await span_element.inner_text() 
            class_attribute = await span_element.get_attribute("class") 

            if text_content == today and ("prevMonthDay" not in (class_attribute or "")) and ("nextMonthDay" not in (class_attribute or "")):
                to_date = span_element
                break  

        if to_date:
            await to_date.click()
        else:
            print("No valid date found!")
        ###################

        await page.click("#rod-submit-type-search")
        await asyncio.sleep(120)

        num_pages_element = page.locator('#rod_type_table_row > div > div div.rod-pages:first-of-type label.rodMxPgLbl')
        num_pages = await num_pages_element.text_content()

        headers = await set_table_headers(page)
        save_to_xlsx(data=None, headers=headers, append=True)

        for i in range(int(num_pages)):
            await scrape_table(page, headers=headers)
            await page.click('#rod_type_table_row > div > div div.rod-pages:first-of-type i.fa-angle-right')

        await browser.close()

ensure_playwright_browsers()
asyncio.run(main())